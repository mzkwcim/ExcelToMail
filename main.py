from openpyxl import load_workbook
import requests

# Ścieżka do pliku Excel
excel_file_path = "C:\\Users\\mzkwcim\\Desktop\\Ośmioboje\\Ośmioboje.xlsx"

# Wczytaj skoroszyt
workbook = load_workbook(filename=excel_file_path)

# Wybierz arkusz, z którego chcesz odczytać dane
sheet = workbook[workbook.sheetnames[0]]
wholeText = ""
# Pobierz dane z komórek
for row in sheet.iter_rows(values_only=True):
    for cell in row:
        wholeText += cell + "\t"
    wholeText += "\n"

# Zamykanie skoroszytu (nieobowiązkowe, ale zalecane)
workbook.close()
print(wholeText)

